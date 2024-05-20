from datetime import datetime
from typing import List

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import reflex as rx
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from plotly.graph_objs import Figure
from scipy.signal import savgol_filter


class FigureWithTitle(rx.Base):
    fig: Figure
    title: str


def authenticate_user(username, password):
    return True


class State(rx.State):
    """The app state."""

    figs1: List[FigureWithTitle] = []
    figs2: List[FigureWithTitle] = []
    total: FigureWithTitle = None
    loaded_document: pd.DataFrame = pd.DataFrame()
    predicted_document: pd.DataFrame = pd.DataFrame()
    date: str = ''
    best_model: str = 'Arima'
    best_score: str = 'SMAPE = 6.7%, MSE = 450.787'
    is_authenticated: bool = False
    print(datetime.now())

    async def handle_upload(self, files: list[rx.UploadFile]):
        """Handle the upload of file(s).

        Args:
            files: The uploaded files.
        """
        for file in files:
            upload_data = await file.read()
            outfile = rx.get_upload_dir() / file.filename

            with outfile.open("wb") as file_object:
                file_object.write(upload_data)

            self.parse_loaded_file(outfile)

    def parse_formula(self, test=True, formula_str=''):
        if test:
            return lambda x: x[0] * x[1] + x[2] * x[3] + x[4] * x[5]
        tokens = Tokenizer(formula_str).items
        formula_tokens = [t.value for t in tokens if t.type == 'operand']
        # TODO cast to Python function
        formula = ''.join(formula_tokens)
        return formula

    def predict(self, test=True):
        if test:
            self.predicted_document = pd.read_excel('data/test_data_output.xlsx', index_col=0)
            self.predicted_document.index = pd.to_datetime(self.predicted_document.index)
        else:
            # TODO add actual predictions
            self.predicted_document = pd.read_excel('data/test_data_output.xlsx', index_col=0)
            self.predicted_document.index = pd.to_datetime(self.predicted_document.index)
        self.predicted_document['total'] = self.predicted_document.apply(self.formula, axis=1)
        return self.predicted_document

    def plot_with_conf(self, col) -> Figure:
        df = self.predicted_document
        CI = np.quantile(df[col], 0.01)
        fig = go.Figure()

        fig.add_traces(
            [
                go.Scatter(
                    x=df[df.index < self.date].index,
                    y=df[df.index < self.date][col].values,
                    mode="lines",
                    name='history'
                ),
                go.Scatter(
                    x=df[df.index >= self.date].index,
                    y=df[df.index >= self.date][col].values,
                    mode="lines",
                    name='forecast'
                ),
                go.Scatter(
                    x=df[df.index >= self.date].index,
                    y=df[df.index >= self.date][col] + savgol_filter(
                        np.random.random(size=df[df.index >= self.date].shape[0]),
                        50, 2
                    ) * CI,
                    mode='lines',
                    line_color='rgba(0,0,0,0)',
                    showlegend=False
                ),
                go.Scatter(
                    x=df[df.index >= self.date].index,
                    y=df[df.index >= self.date][col] - savgol_filter(
                        np.random.random(size=df[df.index >= self.date].shape[0]),
                        50,
                        2
                    ) * CI,
                    mode='lines',
                    line_color='rgba(0,0,0,0)',
                    name='95% confidence interval',
                    fill='tonexty',
                    fillcolor='rgba(255, 0, 0, 0.2)'
                )
            ]
        )
        return fig

    def plot_finals(self):
        self.predict()
        for cols in np.array(self.predicted_document.columns[:6]).reshape((3, 2)).tolist():
            figs = []
            for col in cols:
                fig = self.plot_with_conf(col=col)
                figs.append(FigureWithTitle(fig=fig, title=col))
            self.figs1.append(figs[0])
            self.figs2.append(figs[1])

        if 'total' in self.predicted_document.columns:
            self.total = FigureWithTitle(
                fig=self.plot_with_conf(col='total'),
                title='Total'
            )
            self.total.fig.write_image(f"assets/fig.png")

    def parse_loaded_file(self, outfile):
        df = pd.read_excel(outfile, index_col=0)
        df.index = pd.to_datetime(df.index)

        if 'total' in df.columns:
            wb = load_workbook(outfile)
            sheet = wb.active
            # Прочитать формулу из последней строки и последнего столбца
            formula_str = sheet.cell(row=2, column=sheet.max_column).value
            self.parse_formula(formula_str=formula_str)
        else:
            self.parse_formula()

        # TODO validation
        self.loaded_document = df
        self.date = str(df[df.isna().any(axis=1)].index.min())[:10]
        if 'total' in self.loaded_document.columns:
            self.formula = self.parse_formula()

    def login_submit(self, form):
        # rx.redirect('/upload')

        self.login_error_message = ""
        username = form["username"]
        password = form["password"]
        is_authenticated = authenticate_user(username, password)

        if not is_authenticated:
            self.login_error_message = "He удалось войти в CIBAA LLM UI"
            return rx.set_value("password", "")

        # TODO validation
        self.is_authenticated = True
        return rx.redirect('/upload')

    def redirect(self) -> rx.Component:
        if not self.user.is_authenticated:
            return rx.box(rx.redirect("/"))
        return rx.box(rx.redirect("/upload"))

    def clean(self):
        self.loaded_document = pd.DataFrame()
        self.date = ''
        self.total = None
        self.figs1.clear()
        self.figs2.clear()

    def check_authentication(self):
        print(self.is_authenticated)
        return rx.cond(
            self.is_authenticated & True,
            rx.fragment(),
            self.redirect()
        )
